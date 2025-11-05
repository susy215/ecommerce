"""
Generador de reportes en múltiples formatos (PDF, Excel, CSV)
"""
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import csv


class GeneradorReportes:
    """
    Genera reportes en diferentes formatos basados en los datos de consulta
    """
    
    def __init__(self, datos_consulta, interpretacion):
        self.datos = datos_consulta
        self.params = interpretacion
        self.titulo = self._generar_titulo()
    
    def _generar_titulo(self):
        """Genera un título descriptivo para el reporte"""
        tipo = self.params['tipo_reporte']
        # Título más amigable por tipo
        if tipo == 'inventario':
            titulo = "Inventario actual"
        elif tipo == 'ventas':
            titulo = "Reporte de Ventas"
        elif tipo == 'clientes':
            titulo = "Reporte de Clientes"
        elif tipo == 'productos':
            titulo = "Reporte de Productos"
        else:
            titulo = f"Reporte de {tipo.title()}"
        
        if self.params['fecha_inicio'] and self.params['fecha_fin'] and tipo != 'inventario':
            inicio = self.params['fecha_inicio'].strftime('%d/%m/%Y')
            fin = self.params['fecha_fin'].strftime('%d/%m/%Y')
            titulo += f" del {inicio} al {fin}"
        elif self.params['fecha_inicio'] and tipo != 'inventario':
            inicio = self.params['fecha_inicio'].strftime('%d/%m/%Y')
            titulo += f" desde {inicio}"
        
        if self.params['agrupar_por']:
            agrupaciones = ', '.join(self.params['agrupar_por'])
            titulo += f" - Agrupado por {agrupaciones}"
        
        return titulo
    
    def generar(self, formato='pdf'):
        """Genera el reporte en el formato especificado"""
        if formato == 'pdf':
            return self.generar_pdf()
        elif formato == 'excel':
            return self.generar_excel()
        elif formato == 'csv':
            return self.generar_csv()
        else:
            return None
    
    def generar_pdf(self):
        """Genera reporte en PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Contenedor de elementos
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Estilo para subtítulo
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#7f8c8d'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Título
        elements.append(Paragraph(self.titulo, title_style))
        
        # Fecha de generación
        fecha_gen = timezone.now().strftime('%d/%m/%Y %H:%M')
        elements.append(Paragraph(f"Generado el: {fecha_gen}", subtitle_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Verificar si hay datos
        if not self.datos.get('datos'):
            elements.append(Paragraph("No se encontraron datos para este reporte", styles['Normal']))
            doc.build(elements)
            buffer.seek(0)
            return buffer
        
        # Crear tabla
        columnas = self.datos.get('columnas', [])
        datos_tabla = []
        
        # Encabezados
        headers = [col.replace('_', ' ').title() for col in columnas]
        datos_tabla.append(headers)
        
        # Datos
        for fila in self.datos['datos']:
            if isinstance(fila, dict):
                row = [str(fila.get(col, '')) for col in columnas]
                datos_tabla.append(row)
        
        # Crear tabla
        tabla = Table(datos_tabla, repeatRows=1)
        
        # Estilo de tabla
        tabla.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Cuerpo
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Líneas
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2980b9')),
            
            # Filas alternas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')]),
        ]))
        
        elements.append(tabla)
        
        # Resumen si hay métricas
        if len(self.datos['datos']) > 0:
            elements.append(Spacer(1, 0.3*inch))
            resumen_style = ParagraphStyle(
                'Resumen',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=6,
                fontName='Helvetica-Bold'
            )
            elements.append(Paragraph(f"Total de registros: {len(self.datos['datos'])}", resumen_style))
            # Resumen detallado para inventario
            if self.params.get('tipo_reporte') == 'inventario':
                try:
                    total_stock = sum(int(f.get('stock', 0)) for f in self.datos['datos'])
                    total_valor = sum(float(f.get('valor_inventario', 0)) for f in self.datos['datos'])
                    elements.append(Paragraph(f"Stock total: {total_stock}", styles['Normal']))
                    elements.append(Paragraph(f"Valor total de inventario: ${total_valor:,.2f}", styles['Normal']))
                except Exception:
                    pass
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generar_excel(self):
        """Genera reporte en Excel"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border_side = Side(border_style="thin", color="000000")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Título
        sheet.merge_cells('A1:' + chr(65 + len(self.datos.get('columnas', [])) - 1) + '1')
        title_cell = sheet['A1']
        title_cell.value = self.titulo
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Fecha de generación
        sheet.merge_cells('A2:' + chr(65 + len(self.datos.get('columnas', [])) - 1) + '2')
        fecha_cell = sheet['A2']
        fecha_cell.value = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        fecha_cell.alignment = Alignment(horizontal="center")
        
        # Encabezados (fila 4)
        columnas = self.datos.get('columnas', [])
        for col_idx, col_name in enumerate(columnas, start=1):
            cell = sheet.cell(row=4, column=col_idx)
            cell.value = col_name.replace('_', ' ').title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        if self.datos.get('datos'):
            for row_idx, fila in enumerate(self.datos['datos'], start=5):
                for col_idx, col_name in enumerate(columnas, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    valor = fila.get(col_name, '')
                    cell.value = valor
                    cell.border = border
                    
                    # Alineación de números
                    if isinstance(valor, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(columnas) + 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = 20
        
        # Guardar en buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generar_csv(self):
        """Genera reporte en CSV"""
        buffer = BytesIO()
        
        # Escribir en modo texto
        import io
        text_buffer = io.StringIO()
        
        columnas = self.datos.get('columnas', [])
        writer = csv.writer(text_buffer, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        
        # Título
        writer.writerow([self.titulo])
        writer.writerow([f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"])
        writer.writerow([])  # Línea vacía
        
        # Encabezados
        headers = [col.replace('_', ' ').title() for col in columnas]
        writer.writerow(headers)
        
        # Datos
        if self.datos.get('datos'):
            for fila in self.datos['datos']:
                row = [fila.get(col, '') for col in columnas]
                writer.writerow(row)
        
        # Convertir a bytes
        buffer.write(text_buffer.getvalue().encode('utf-8-sig'))  # BOM para Excel
        buffer.seek(0)
        return buffer
